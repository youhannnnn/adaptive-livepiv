import sys
import os
import threading
import cv2
from PyQt5.QtWidgets import (QApplication, QWidget, QVBoxLayout, QHBoxLayout, QDialog,
                             QPushButton, QListWidget, QLabel, QFileDialog, QComboBox,
                             QMessageBox, QSizePolicy, QSplitter, QDialogButtonBox,
                             QLineEdit, QFormLayout, QRadioButton, QStackedWidget)
from PyQt5.QtCore import Qt, QDir, QPoint, QRect, QUrl, QTimer
from PyQt5.QtGui import QPixmap, QPainter, QPen, QColor, QCursor, QImage, QIcon
import beckend
import glob
import pandas as pd
import livePIV_DrawLabel
from openpyxl import load_workbook
from PIL import Image
import PlotAverageVorticityCloudMap
import PlotAverageVelocityStreamline
import PlotFlowPulsationIntensityCloudMap
import PlotInstantaneousVorticityCloudMap


# 图像参数设置对话框
class ImageParamsDialog(QDialog):
    def __init__(self, parent=None):  # 构造函数，初始化对话框，可以接受一个父窗口对象
        super().__init__(parent)  # 调用父类的构造方法，确保对话框正确初始化
        self.setWindowTitle("图像参数设置")  # 设置对话框标题为“图像参数设置”
        layout = QVBoxLayout(self)  # 创建一个垂直布局管理器，并将其设为当前对话框的主布局
        # 创建表单布局
        form_layout = QFormLayout()  # 创建一个表单布局对象（label + 控件形式）
        self.pixels_input = QLineEdit()  # 创建一个文本输入框，用于输入“像素每米”的数值
        self.time_input = QLineEdit()  # 创建一个文本输入框，用于输入“时间差”的数值
        form_layout.addRow("像素每米（px/m）：", self.pixels_input)  # 向表单中添加一行标签和输入框，提示用户输入像素密度
        form_layout.addRow("时间差（s）：", self.time_input)  # 添加第二行标签和输入框，提示用户输入帧之间的时间差
        layout.addLayout(form_layout)  # 将表单布局添加到主垂直布局中

        # 按钮组
        self.btn_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)  # 创建标准的对话框按钮组，包含“确定”和“取消”按钮
        layout.addWidget(self.btn_box)  # 将按钮组添加到主布局中
        self.btn_box.accepted.connect(self.accept)  # 将“确定”按钮的点击事件连接到QDialog自带的accept()槽函数（关闭对话框并返回Accepted）
        self.btn_box.rejected.connect(self.reject)  # 将“取消”按钮的点击事件连接到reject()槽函数（关闭对话框并返回Rejected）

    def get_params(self):                       # 定义一个公共方法，用于获取用户输入的两个参数
        """返回输入参数的元组（像素每米, 时间差）"""  # 方法说明文档，说明返回值为一个包含两个文本值的元组
        return (self.pixels_input.text(), self.time_input.text())  # 获取两个输入框中的文本并返回为元组


# 掩膜设置对话框
class MaskDialog(QDialog):  # 点击掩膜后设置掩膜的对话框，继承自 QDialog（对话框类）
    def __init__(self, parent=None):  # 构造函数，接受一个可选的父窗口对象
        super().__init__(parent)  # 调用父类构造函数，初始化对话框
        self.setWindowTitle("掩膜设置")  # 设置对话框窗口的标题为“掩膜设置”
        self.layout = QVBoxLayout(self)  # 创建一个垂直布局管理器作为主布局，并设置给该对话框

        # 形状选择
        self.shape_combo = QComboBox()  # 创建一个下拉框组件，用于选择掩膜的形状（例如矩形、圆形）
        self.shape_combo.addItems(["矩形", "圆形"])  # 将可选项“矩形”和“圆形”添加到下拉框中
        self.layout.addWidget(self.shape_combo)  # 将下拉框控件添加到垂直布局中
        # 操作按钮
        self.btn_box = QDialogButtonBox()  # 创建一个按钮盒子（用于容纳多个按钮），这里不使用标准按钮类型
        self.btn_apply = QPushButton("执行掩膜")  # 创建一个普通按钮，文本为“执行掩膜”
        self.btn_clear = QPushButton("清除掩膜")  # 创建另一个普通按钮，文本为“清除掩膜”
        self.btn_box.addButton(self.btn_apply, QDialogButtonBox.ActionRole)  # 将“执行掩膜”按钮添加到按钮盒子中，指定角色为自定义操作
        self.btn_box.addButton(self.btn_clear, QDialogButtonBox.ActionRole)  # 将“清除掩膜”按钮也添加到按钮盒子中
        self.layout.addWidget(self.btn_box)  # 将按钮盒子添加到垂直主布局中


# 主窗口类
class ImageBrowser(QWidget):  # 定义一个图像浏览器类，继承自 QWidget，用于构建主应用窗口
    def __init__(self):  # 构造函数，初始化主窗口及相关组件
        super().__init__()      # 调用 QWidget 的父类构造函数
        self.init_parameter()    # 初始化跟后端连接需要传递的各个参数
        self.init_ui()           # 调用初始化界面方法，创建布局与控件

    def init_parameter(self):
        self.image_paths = []  # 初始化图像路径列表，用于保存所有导入图像的完整路径
        self.output_dir = None  # 初始化输出目录为 None，用于后续保存输出路径
        self.result_path = None         # 结果文件夹
        self.result_figure_name = None      # 结果图像名称
        self.cache_path = None      # 缓存路径
        self.mask_shapes = []  # 存储所有掩膜图形的列表（包含类型
        self.drawing = False  # 是否处于绘图模式和矩形位置）
        self.current_shape = None  # 当前正在绘制的掩膜形状（暂时未使用，可扩展）（备用变量）
        self.drag_mode = False  # 是否处于拖动图形模式
        self.selected_shape_index = -1  # 当前选中的图形索引（默认无选中）
        self.pixels_per_meter = None  # 初始化像素密度参数（单位 px/m）
        self.time_interval = None  # 初始化时间间隔参数（单位 s）
        self.display_area = []  # 显示框的宽度和高度
        self.cv_capture = None  # OpenCV视频捕获对象
        self.cv_timer = QTimer()  # 用于更新视频帧的定时器
        self.last_frame_of_video = None  # 视频的最后一帧
        self.video_url = "http://100.78.147.151:8080/video"
        self.is_recording = False   # 录制状态
        self.video_writer = None    # 视频写入对象
        self.recording_fps = 30     # 录制视频的帧率

    def init_ui(self):  # 初始化主界面的方法，设置所有控件和布局
        self.setWindowTitle('实时PIV测量系统')  # 设置窗口标题
        self.setGeometry(100, 100, 1000, 600)  # 设置窗口初始位置和大小（x=100, y=100, 宽=1000，高=600）

        # 左侧面板
        left_panel = QWidget()  # 创建左侧控制面板容器
        left_layout = QVBoxLayout(left_panel)  # 为左侧面板创建垂直布局
        left_layout.setContentsMargins(5, 5, 5, 5)  # 设置边距为5
        left_layout.setSpacing(10)  # 设置控件之间的垂直间距为10像素

        # 右侧面板
        right_container = QWidget()  # 创建右侧容器控件，用于包含图像显示区与按钮面板
        right_layout = QVBoxLayout(right_container)  # 创建垂直布局用于右侧容器
        right_layout.setContentsMargins(0, 0, 0, 0)  # 设置布局边距为0
        right_layout.setSpacing(5)  # 设置控件之间的间距为5像素

        # 主布局使用水平分割器
        main_splitter = QSplitter(Qt.Horizontal)  # 创建水平分割器，将窗口分为左右两部分
        main_splitter.addWidget(left_panel)  # 将左侧控制面板加入水平分割器
        main_splitter.addWidget(right_container)  # 将右侧容器加入水平分割器
        main_splitter.setSizes([250, 750])  # 设置左右分割初始比例：左 250px，右 750px（大约 1:3）

        # 左侧面板的图像列表框
        self.left_list_widget = QListWidget()  # 创建图像文件列表控件
        left_layout.addWidget(self.left_list_widget)  # 将图像列表添加到左侧布局中
        # 左侧面板的按钮
        left_btn_layout = QHBoxLayout()  # 创建一个水平布局，用于排列添加和清空按钮
        self.btn_add = QPushButton("添加图像")  # 创建一个“添加图像”按钮
        self.btn_add.setFixedHeight(30)  # 设置按钮高度为35像素
        self.btn_clear = QPushButton("清空列表")  # 创建一个“清空列表”按钮
        self.btn_clear.setFixedHeight(30)  # 设置按钮高度为35像素
        left_btn_layout.addWidget(self.btn_add)  # 将“添加图像”按钮添加到按钮水平布局中
        left_btn_layout.addWidget(self.btn_clear)  # 将“清空列表”按钮添加到按钮水平布局中
        left_layout.addLayout(left_btn_layout)  # 将按钮布局添加到左侧主布局

        # 右侧显示区域
        self.stacked_display = QStackedWidget()
        self.right_show_panel = livePIV_DrawLabel.DrawLabel(self)
        self.right_show_panel.setAlignment(Qt.AlignCenter)
        self.right_show_panel.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        self.cv_video_widget = QWidget()
        self.cv_video_layout = QVBoxLayout(self.cv_video_widget)
        self.cv_video_label = QLabel()
        self.cv_video_label.setAlignment(Qt.AlignCenter)
        self.cv_video_label.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        self.cv_video_layout.addWidget(self.cv_video_label)
        self.stacked_display.addWidget(self.right_show_panel)  # 索引 0: 图像
        self.stacked_display.addWidget(self.cv_video_widget)  # 索引 1: OpenCV视频
        right_layout.addWidget(self.stacked_display, 1)
        # 右侧面板的按钮
        right_btn_layout = QHBoxLayout()  # 创建一个水平布局用于容纳按钮
        right_btn_layout.setContentsMargins(5, 5, 5, 5)  # 设置内边距为 5 像素
        self.btn_output = QPushButton("输出路径")  # 创建“输出路径”按钮
        self.btn_output.setFixedHeight(30)  # 设置按钮高度为30像素
        self.btn_mask = QPushButton("掩膜操作")  # 创建“掩膜操作”按钮（用于绘制和清除掩膜）
        self.btn_mask.setFixedHeight(30)  # 设置按钮高度为30像素
        self.btn_size = QPushButton("图像尺寸")  # 创建“图像尺寸”按钮（用于设置像素密度与时间差）
        self.btn_size.setFixedHeight(30)  # 设置按钮高度为30像素
        self.btn_process = QPushButton("执行计算")  # 创建“执行计算”按钮（执行分析或演示操作）
        self.btn_process.setFixedHeight(30)  # 设置按钮高度为30像素
        self.btn_live_mode = QPushButton("转入实况")
        self.btn_live_mode.setFixedHeight(30)
        self.btn_record_start = QPushButton("录制视频")
        self.btn_record_start.hide()    # 先隐藏录制按钮
        self.btn_record_start.setFixedHeight(30)
        self.btn_record_stop = QPushButton("停止录制")
        self.btn_record_stop.hide()     # 先隐藏停止录制按钮
        self.btn_record_stop.setFixedHeight(30)
        btn_style = "QPushButton {padding:5px; font-size:12px; min-width:70px;}"  # 设置统一的按钮样式（填充、字体大小、最小宽度）
        self.btn_output.setStyleSheet(btn_style)  # 应用样式
        self.btn_mask.setStyleSheet(btn_style)
        self.btn_size.setStyleSheet(btn_style)
        self.btn_process.setStyleSheet(btn_style)
        self.btn_live_mode.setStyleSheet(btn_style)
        self.btn_record_start.setStyleSheet(btn_style)
        self.btn_record_stop.setStyleSheet(btn_style)
        right_btn_layout.addWidget(self.btn_output)  # 添加“输出路径”按钮
        right_btn_layout.addWidget(self.btn_mask)  # 添加“掩膜操作”按钮
        right_btn_layout.addWidget(self.btn_size)  # 添加“图像尺寸”按钮
        right_btn_layout.addWidget(self.btn_process)  # 添加“执行计算”按钮
        right_btn_layout.addStretch(1)  # 添加弹性空白项，使按钮靠左排列，右侧留白
        right_btn_layout.addWidget(self.btn_record_start, alignment=Qt.AlignRight)
        right_btn_layout.addWidget(self.btn_record_stop, alignment=Qt.AlignRight)
        right_btn_layout.addWidget(self.btn_live_mode, alignment=Qt.AlignRight) # 实况模式按钮在最右侧
        right_layout.addLayout(right_btn_layout)  # 将控制面板添加到底部的右侧垂直布局中

        # 主窗口布局
        main_layout = QHBoxLayout(self)  # 设置主窗口为水平布局（实际上只放一个 main_splitter）
        main_layout.addWidget(main_splitter)  # 添加左右分割器

        # 连接信号
        self.btn_add.clicked.connect(self.show_selection_dialog)  # “添加图像”按钮连接到图像添加对话框方法
        self.btn_clear.clicked.connect(self.clear_list)  # “清空列表”按钮连接到清空方法
        self.left_list_widget.itemClicked.connect(self.on_item_clicked)     # 图像列表项点击时，显示对应图像
        self.btn_output.clicked.connect(self.select_output_dir)  # 连接按钮点击信号到选择输出目录的方法
        self.btn_mask.clicked.connect(self.show_mask_dialog)  # 将“掩膜操作”按钮与 show_mask_dialog 方法连接
        self.btn_size.clicked.connect(self.show_image_params_dialog)  # “图像尺寸”按钮连接到图像参数对话框
        self.btn_process.clicked.connect(self.execute_calculation)  # “执行计算”按钮连接到计算执行方法
        self.btn_live_mode.clicked.connect(self.connect_to_live_mode)
        self.btn_record_start.clicked.connect(self.start_recording)
        self.btn_record_stop.clicked.connect(self.stop_recording)
        self.cv_timer.timeout.connect(self.update_cv_frame)     # 连接定时器到更新函数

    def show_selection_dialog(self):  # 显示添加图像的方式选择对话框（图像对 or 批量添加）
        # 第一步：显示选择对话框
        choice = QMessageBox.question(  # 弹出一个问题对话框，询问用户添加图像的方式
            self,
            '添加图像',
            '可以导入视频或图像\n导入图像的数量必须为偶数',
            QMessageBox.Yes | QMessageBox.Cancel  # 三个选项：是、否、取消
        )
        if choice == QMessageBox.Yes:  # 用户选择“是” → 添加图像对（两帧）
            self.add_image_pair()  # 调用添加图像对方法

    def add_image_pair(self):  # 添加图像对的方法（用户选择两帧图像用于配对处理），选择一对图像 用于之后piv分析
        # 获取初始目录（使用最近访问的目录或默认目录）
        init_dir = QDir.homePath() if not self.image_paths else os.path.dirname(self.image_paths[-1])
        # 选择文件
        files, _ = QFileDialog.getOpenFileNames(  # 打开文件选择对话框，允许多选文件
            self,
            "添加图像",  # 对话框标题
            init_dir,  # 初始路径
            "图像文件 (*.jpg *.jpeg *.png *.tif *.tiff *.bmp *.mp4)"  # 文件类型过滤器
        )
        if len(files) == 1 and files[0].endswith(('.mp4', '.avi')):
            self.clear_list()
        if len(files) > 1 and files[0].endswith(('.mp4', '.avi')):
            QMessageBox.warning(self, "选择错误", "视频文件只能添加一个")
            return
        if files[0].endswith(('.jpg', '.jpeg', '.png', '.tif', '.tiff', '.bmp')) and self.image_paths:
            if self.image_paths[0].lower().endswith(('.mp4', '.avi')):
                self.clear_list()
        self._add_files(files)

    def _add_files(self, files):  # 将文件路径添加到列表中显示，同时保存路径数据
        self.image_paths.extend(files)  # 将新导入的文件路径追加到 image_paths 列表
        self.left_list_widget.addItems([os.path.basename(f) for f in files])  # 在左侧列表中显示每个图像的文件名（不含路径）

    def clear_list(self):  # 清空图像列表的方法
        self.image_paths.clear()  # 清空图像路径列表
        self.left_list_widget.clear()  # 清空左侧图像文件名列表
        self.right_show_panel.clear()  # 清空图像显示区域

    def on_item_clicked(self, item):
        row = self.left_list_widget.row(item)
        self.show_image(row)

    def show_image(self, index):  # 显示当前选中的图像（列表中第 index 项）
        if 0 <= index < len(self.image_paths):  # 检查索引是否有效
            file_path = self.image_paths[index]
            try:
                if file_path.lower().endswith(('.mp4', '.avi')):  # 判断是否为视频文件
                    # 停止之前的视频
                    self.cv_timer.stop()
                    if self.cv_capture is not None:
                        self.cv_capture.release()
                    # 使用OpenCV打开视频
                    self.cv_capture = cv2.VideoCapture(file_path)
                    if not self.cv_capture.isOpened():
                        QMessageBox.critical(self, "错误", "无法打开视频文件")
                        return
                    # 切换到OpenCV视频页面
                    self.stacked_display.setCurrentIndex(1)
                    # 启动定时器，更新帧
                    self.cv_timer.start(30)  # 30毫秒一帧，大约33fps
                else:  # 默认处理方式：加载并显示图片
                    pixmap = QPixmap(file_path)
                    if pixmap.isNull():
                        raise ValueError("无效的图像文件")
                    scaled_pixmap = pixmap.scaled(
                        self.right_show_panel.size(),
                        Qt.KeepAspectRatio,
                        Qt.SmoothTransformation
                    )
                    self.right_show_panel.setPixmap(scaled_pixmap)
                    # 切换回图像页面
                    self.stacked_display.setCurrentIndex(0)
            except Exception as e:
                QMessageBox.critical(self, "错误", f"无法加载文件：\n{str(e)}")

    def update_cv_frame(self):
        # 通过计时器不断更新视频帧，达到播放视频的效果
        if self.cv_capture is not None:
            ret, frame = self.cv_capture.read()
            if ret:
                # 保存当前帧（BGR格式）
                self.last_frame_of_video = frame.copy()

                # 如果正在录制，写入当前帧
                if self.is_recording and self.video_writer is not None:
                    # 确保帧大小与视频写入器一致
                    if frame.shape[:2] != (self.recording_height, self.recording_width):
                        frame = cv2.resize(frame, (self.recording_width, self.recording_height))
                    self.video_writer.write(frame)  # 直接写入BGR格式

                # 转换BGR到RGB用于显示
                frame_rgb = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
                # 创建QImage
                h, w, ch = frame_rgb.shape
                bytes_per_line = ch * w
                qt_image = QImage(frame_rgb.data, w, h, bytes_per_line, QImage.Format_RGB888)
                # 缩放并显示
                scaled_pixmap = QPixmap.fromImage(qt_image).scaled(
                    self.cv_video_label.size(),
                    Qt.KeepAspectRatio,
                    Qt.SmoothTransformation
                )
                self.cv_video_label.setPixmap(scaled_pixmap)
            else:
                if self.btn_live_mode.text() == "退出实况":
                    # 视频流中断，退出实况模式
                    QMessageBox.warning(self, "连接中断", "与IP摄像头的连接已中断")
                    self.exit_the_live_mode()
                else:
                    # 视频播放完毕，停止定时器
                    self.cv_timer.stop()
                    # 如果有最后一帧，显示在DrawLabel上
                    if self.last_frame_of_video is not None:
                        self.display_last_frame()
                        # 切换到图像页面以便进行掩膜操作
                        self.stacked_display.setCurrentIndex(0)

    def display_last_frame(self):
        """将视频的最后一帧显示在DrawLabel上"""
        if self.last_frame_of_video is not None:
            # 转换BGR到RGB
            frame_rgb = cv2.cvtColor(self.last_frame_of_video, cv2.COLOR_BGR2RGB)
            # 创建QImage
            h, w, ch = frame_rgb.shape
            bytes_per_line = ch * w
            qt_image = QImage(frame_rgb.data, w, h, bytes_per_line, QImage.Format_RGB888)
            # 转换为QPixmap
            pixmap = QPixmap.fromImage(qt_image)
            # 缩放并显示在DrawLabel上
            scaled_pixmap = pixmap.scaled(
                self.right_show_panel.size(),
                Qt.KeepAspectRatio,
                Qt.SmoothTransformation
            )
            self.right_show_panel.setPixmap(scaled_pixmap)

    def closeEvent(self, event):
        # 关闭窗口时停止定时器并释放视频捕获
        self.cv_timer.stop()
        if self.cv_capture is not None:
            self.cv_capture.release()
        super().closeEvent(event)

    def select_output_dir(self):  # 选择输出路径的对话框
        """选择输出路径"""
        dir_path = QFileDialog.getExistingDirectory(  # 打开文件夹选择对话框
            self,
            "选择输出文件夹",
            QDir.homePath(),  # 默认路径为用户主目录
            QFileDialog.ShowDirsOnly  # 仅允许选择文件夹
        )
        if dir_path:  # 如果用户选择了路径
            self.output_dir = dir_path  # 保存路径
            QMessageBox.information(  # 弹出提示框确认路径设置成功
                self,
                "路径设置成功",
                f"输出路径已设置为：\n{dir_path}",
                QMessageBox.Ok
            )

    def show_mask_dialog(self):  # 显示掩膜设置对话框（选择图形 + 清除）
        dialog = MaskDialog(self)  # 创建掩膜对话框对象，主窗口为父窗口
        dialog.btn_apply.clicked.connect(  # 连接“执行掩膜”按钮点击事件
            lambda: self._handle_apply_mask(dialog)  # 使用 lambda 表达式调用处理函数并传入当前对话框
        )
        dialog.btn_clear.clicked.connect(self.clear_masks)  # 连接“清除掩膜”按钮点击事件
        dialog.exec_()  # 显示对话框，阻塞主界面，等待用户操作

    def show_image_params_dialog(self):  # 显示图像参数设置对话框的方法
        """显示图像参数设置对话框"""  # 方法说明：用于打开像素密度和时间间隔设置窗口
        dialog = ImageParamsDialog(self)  # 创建图像参数设置对话框对象，父窗口为当前主窗口
        if dialog.exec_() == QDialog.Accepted:  # 执行对话框并等待用户操作，若点击“确定”
            px_str, time_str = dialog.get_params()  # 从对话框中获取两个输入值（返回字符串）
            try:
                self.pixels_per_meter = float(px_str)  # 将输入的像素密度转换为浮点数
                self.time_interval = float(time_str)  # 将输入的时间间隔转换为浮点数
                msg_box = QMessageBox(self)
                msg_box.setWindowTitle("参数已保存")
                msg_box.setText(f"参数设置成功：\n像素每米 = {self.pixels_per_meter} px/m\n时间间隔 = {self.time_interval} s")
                msg_box.setStandardButtons(QMessageBox.Ok)
                # 移除默认图标
                msg_box.setIcon(QMessageBox.NoIcon)
                # 设置样式表，使背景与OK按钮一致
                msg_box.setStyleSheet("""
                    QMessageBox {
                        background-color: #f0f0f0;  /* 与OK按钮相似的背景色 */
                    }
                    QMessageBox QLabel {
                        color: #000000;  /* 文本颜色 */
                    }
                """)
                msg_box.exec_()
            except ValueError:  # 若输入无法转换为数字（非法字符）
                QMessageBox.critical(  # 弹出错误提示框
                    self,
                    "输入错误",
                    "请输入有效的数字参数！",
                    QMessageBox.Ok
                )

    def execute_calculation(self):  # 执行计算的方法，受“执行计算”按钮触发
        """执行计算按钮功能"""
        if self.pixels_per_meter is None or self.time_interval is None:
            QMessageBox.warning(self, "警告", "请先设置图像参数！")
            return
        if self.output_dir is None:
            QMessageBox.warning(self, "警告", "请先设置输出路径！")
            return
        if self.image_paths and self.image_paths[0].lower().endswith(('.mp4', '.avi')):
            self.execute_video_calculation()
        if self.image_paths and self.image_paths[0].lower().endswith(('.jpg', '.jpeg', '.png', '.tif', '.tiff', '.bmp')):
            if len(self.image_paths) % 2 == 0:
                self.execute_image_calculation()
            else:
                QMessageBox.warning(self, "数据量错误", "请导入偶数张图像！")
        QMessageBox.information(
            self,
            "计算完成",
            "执行成功！\n\n（当前为演示功能）",
            QMessageBox.Ok
        )

    def execute_video_calculation(self):
        if self.cache_path is None:
            cache_path = QFileDialog.getExistingDirectory(
                self,
                "选择缓存路径",
                QDir.homePath(),
                QFileDialog.ShowDirsOnly
            )
            if cache_path:
                self.cache_path = cache_path
            else:
                QMessageBox.warning(self, "警告", "请先选择缓存路径！")
                return
        # 获取视频路径
        video_path = self.image_paths[0]
        # 打开视频文件
        cap = cv2.VideoCapture(video_path)
        if not cap.isOpened():
            QMessageBox.critical(self, "错误", "无法打开视频文件！")
            return
        # 获取视频信息
        total_frames = int(cap.get(cv2.CAP_PROP_FRAME_COUNT))
        # --------------------------------------------------------------这里后面改成少于20帧也能分析
        if total_frames < 20:
            QMessageBox.warning(self, "警告", "视频帧数不足20帧，无法进行分析！")
            cap.release()
            return
        # 计算帧间隔
        frame_interval = max(1, total_frames // 9)  # 确保至少提取20帧
        # 确定要提取的帧索引
        frame_indices = []
        for i in range(10):
            if i == 0:
                frame_indices.append(0)  # 第一帧
                frame_indices.append(1)
            else:
                frame_index = min(i * frame_interval, total_frames - 2)
                frame_indices.append(frame_index)
                frame_indices.append(frame_index + 1)
        # 提取并保存帧
        frame_files = []
        num = 0
        for i, frame_idx in enumerate(frame_indices):
            cap.set(cv2.CAP_PROP_POS_FRAMES, frame_idx)
            ret, frame = cap.read()
            if ret:
                # 转换为RGB
                frame_rgb = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
                # 保存为图像文件
                frame_filename = f"frame_{i + 1:02d}.png"
                frame_path = os.path.join(self.cache_path, frame_filename)
                cv2.imwrite(frame_path, frame_rgb)
                frame_files.append(frame_path)
                num += 1
        # QMessageBox.information(self, "帧提取完成", f"已提取{num}帧图像作为视频分析数据", QMessageBox.Ok)
        cap.release()
        self.image_paths = frame_files
        # 处理帧对
        for i in range(0, len(frame_files), 2):
            if i + 1 >= len(frame_files):
                break
            frame_a_path = frame_files[i]
            frame_b_path = frame_files[i + 1]
            # 提取帧对的基本信息
            image_path = os.path.dirname(frame_a_path)
            frame_a = os.path.basename(frame_a_path)
            frame_b = os.path.basename(frame_b_path)
            scaling_factor = self.pixels_per_meter * self.time_interval
            self.display_area = [self.right_show_panel.width(), self.right_show_panel.height()]
            # 调用 backend 计算
            self.result_path, self.result_figure_name = beckend.beckend_calculation(
                image_path, self.output_dir, frame_a, frame_b, scaling_factor,
                self.mask_shapes, self.display_area, cloud_chart=True
            )
            # 结果显示并导入excel
            self.show_result_figure()
            self.output_to_excel()
        self.rebuild_excel(len(frame_files)/2)
        self.plot_result_figure()
        # 视频处理完毕后，删除缓存文件，将前后端内容重置为初始状态
        self.clear_masks()
        self.image_paths = video_path
        # -------------------------------------------------------------------------------------------------这里之后要补充清理缓存

    def plot_result_figure(self):
        if self.mask_shapes is None:
            # -------------------------------------------------------------------------------------------这里之后要补充没有掩膜时的自动生成图像处理，现在没写是因为函数形参用0表示的话，会出错，所以暂时不写这里的代码
            return
        frame_a = Image.open(self.image_paths[0])
        figure_width, figure_height = frame_a.size
        scaling_factor = self.pixels_per_meter * self.time_interval
        if self.mask_shapes[0]["type"] == "圆形":
            rect = self.mask_shapes[0]["rect"]
            # 获取圆形掩膜在显示区域中的信息
            x1 = rect.center().x()  # 圆心x
            y1 = rect.center().y()  # 圆心y
            radius_width = rect.width() / 2  # 水平半径
            radius_height = rect.height() / 2  # 垂直半径
            display_width = self.display_area[0]
            display_height = self.display_area[1]
            if figure_width / figure_height > display_width / display_height:
                text_x1 = x1 / display_width * figure_width / scaling_factor / 30 * 100
                text_y1 = figure_height / scaling_factor / 30 * 100 - (y1 - (
                            display_height - figure_height * display_width / figure_width) / 2) / figure_height * figure_width / display_width * figure_height / scaling_factor / 30 * 100
                text_radius_width = radius_width / display_width * figure_width / scaling_factor / 30 * 100
                text_radius_height = radius_height / display_width * figure_width / scaling_factor / 30 * 100
            else:
                text_x1 = (x1 - (display_width - figure_width * display_height / figure_height) / 2) / figure_width * figure_height / display_height * figure_width / scaling_factor / 30 * 100
                text_y1 = figure_height / scaling_factor / 30 * 100 - y1 / display_height * figure_height / scaling_factor / 30 * 100  # y轴调换过来了，所以要前面用一个总长来减
                text_radius_width = radius_width / display_height * figure_height / scaling_factor / 30 * 100
                text_radius_height = radius_height / display_height * figure_height / scaling_factor / 30 * 100
            PlotAverageVorticityCloudMap.plot_vorticity_cloud_cylinder(text_x1, text_y1, max(text_radius_width*2, text_radius_height*2), self.result_path)
            PlotAverageVelocityStreamline.plot_average_velocity_streamline_cylinder(text_x1, text_y1, max(text_radius_width*2, text_radius_height*2), self.result_path)
            PlotFlowPulsationIntensityCloudMap.plot_flow_pulsation_intensity_cloud_map_cylinder(text_x1, text_y1, max(text_radius_width*2, text_radius_height*2), self.result_path)
            PlotInstantaneousVorticityCloudMap.plot_velocity_vector_cylinder(text_x1, text_y1, max(text_radius_width*2, text_radius_height*2), self.result_path)
        if self.mask_shapes[0]["type"] == "矩形":
            rect = self.mask_shapes[0]["rect"]
            top_left = rect.topLeft()  # QPoint
            bottom_right = rect.bottomRight()  # QPoint
            x1, y1 = top_left.x(), top_left.y()
            x2, y2 = bottom_right.x(), bottom_right.y()
            display_width = self.display_area[0]
            display_height = self.display_area[1]
            if figure_width / figure_height > display_width / display_height:
                text_x1 = x1 / display_width * figure_width / scaling_factor / 30 * 100  # （*figure_width/scaling_factor）前面的是相对图片的xy长，乘以这个数之后得到text文件中的xy
                text_x2 = x2 / display_width * figure_width / scaling_factor / 30 * 100
                text_y1 = figure_height / scaling_factor / 30 * 100 - (y1 - (display_height - figure_height * display_width / figure_width) / 2) / figure_height * figure_width / display_width * figure_height / scaling_factor / 30 * 100
                text_y2 = figure_height / scaling_factor / 30 * 100 - (y2 - (display_height - figure_height * display_width / figure_width) / 2) / figure_height * figure_width / display_width * figure_height / scaling_factor / 30 * 100
            else:
                text_x1 = (x1 - (display_width - figure_width * display_height / figure_height) / 2) / figure_width * figure_height / display_height * figure_width / scaling_factor / 30 * 100
                text_x2 = (x2 - (display_width - figure_width * display_height / figure_height) / 2) / figure_width * figure_height / display_height * figure_width / scaling_factor / 30 * 100
                text_y1 = figure_height / scaling_factor / 30 * 100 - y1 / display_height * figure_height / scaling_factor / 30 * 100  # y轴调换过来了，所以要前面用一个总长来减
                text_y2 = figure_height / scaling_factor / 30 * 100 - y2 / display_height * figure_height / scaling_factor / 30 * 100
            PlotAverageVorticityCloudMap.plot_vorticity_cloud_rectangle(text_x1+(text_x2-text_x1)/2, text_y2+(text_y1-text_y2)/2, (text_x2-text_x1), (text_y1-text_y2), self.result_path)
            PlotAverageVelocityStreamline.plot_average_velocity_streamline_rectangle(text_x1+(text_x2-text_x1)/2, text_y2+(text_y1-text_y2)/2, (text_x2-text_x1), (text_y1-text_y2), self.result_path)
            PlotFlowPulsationIntensityCloudMap.plot_flow_pulsation_intensity_cloud_map_rectangle(text_x1+(text_x2-text_x1)/2, text_y2+(text_y1-text_y2)/2, (text_x2-text_x1), (text_y1-text_y2), self.result_path)

    def execute_image_calculation(self):
        for i in range(0, len(self.image_paths), 2):
            image_path = os.path.dirname(self.image_paths[i])
            frame_a = os.path.basename(self.image_paths[i])
            frame_b = os.path.basename(self.image_paths[i + 1])
            scaling_factor = self.pixels_per_meter * self.time_interval
            self.display_area = [self.right_show_panel.width(), self.right_show_panel.height()]
            # 调用 backend 计算
            self.result_path, self.result_figure_name = beckend.beckend_calculation(
                image_path, self.output_dir, frame_a, frame_b, scaling_factor,
                self.mask_shapes, self.display_area, cloud_chart=True
            )
            self.show_result_figure()
            # 将txt文件中的数据导入excel
            self.output_to_excel()
        self.reset_excel_header(len(self.image_paths)/2)

    def show_result_figure(self):
        image_path = os.path.join(self.result_path, self.result_figure_name)
        pixmap = QPixmap(image_path)  # 将结果加载为 QPixmap 对象
        if pixmap.isNull():  # 如果图像无效或加载失败
            print(f"Error: Unable to load image from path: {image_path}")
            raise ValueError("无效的图像文件")  # 抛出异常
        # 自适应显示
        scaled_pixmap = pixmap.scaled(  # 将图像缩放至适配窗口大小
            self.right_show_panel.size(),  # 缩放目标为右侧面板大小
            Qt.KeepAspectRatio,  # 保持原始宽高比
            Qt.SmoothTransformation  # 平滑缩放，提高质量
        )
        self.right_show_panel.setPixmap(scaled_pixmap)  # 设置缩放后的图像到右侧显示区域

    def output_to_excel(self):
        # 步骤1：查找包含 "Open_PIV_results" 的文件夹
        search_path = os.path.join(self.output_dir, 'Open_PIV_results*')
        result_folders = glob.glob(search_path)
        if not result_folders:
            print("未找到以 'Open_PIV_results' 开头的文件夹")
            return
        # 假设我们只处理找到的第一个符合条件的文件夹
        target_folder = result_folders[0]
        # 步骤2：查找目标文件夹内的 field_A000.txt 文件
        txt_file_path = os.path.join(target_folder, 'field_A000.txt')
        if not os.path.exists(txt_file_path):
            print(f"未找到 {txt_file_path} 文件")
            return
        # 读取 txt 文件的数据
        data = []
        with open(txt_file_path, 'r') as file:
            next(file)  # 跳过第一行
            for line in file:
                values = line.split()
                if len(values) >= 4:  # 确保至少有四个值可以提取
                    data.append([float(v) for v in values[:4]])  # 只提取前四个数据 x, y, u, v
        # 步骤3：准备写入 Excel 文件
        excel_file_path = os.path.join(self.output_dir, '分析结果.xlsx')
        df_new = pd.DataFrame(data, columns=['x', 'y', 'u', 'v'])
        try:
            # 如果文件已存在，则加载并找到下一个空列
            df_existing = pd.read_excel(excel_file_path, sheet_name='Sheet1', engine='openpyxl')
            start_col = len(df_existing.columns) + (1 if len(df_existing.columns) % 5 == 0 else 0) + 1 # 计算起始列位置
        except FileNotFoundError:
            # 文件不存在时创建新文件
            df_existing = pd.DataFrame()
            start_col = 0  # 从第0列开始
        with pd.ExcelWriter(excel_file_path, mode='w', engine='openpyxl') as writer:
            # 如果文件已经存在，则先写入现有数据
            if not df_existing.empty:
                df_existing.to_excel(writer, sheet_name='Sheet1', index=False, startrow=0, startcol=0)
            # 写入新的数据
            df_new.to_excel(writer, sheet_name='Sheet1', index=False, startrow=0, startcol=start_col, header=True)
            # 添加表头
            workbook = writer.book
            worksheet = writer.sheets['Sheet1']
            for col_num, value in enumerate(['x', 'y', 'u', 'v'], start=start_col):
                worksheet.cell(row=1, column=col_num + 1, value=value)

    def rebuild_excel(self, frame_pair_num):
        # 路径定义
        excel_file_path = os.path.join(self.output_dir, '分析结果.xlsx')
        avg_excel_path = os.path.join(self.output_dir, '平均结果.xlsx')
        # 表头模板：每组 x y u v 空
        header_group = ['x', 'y', 'u', 'v', '']
        full_header = header_group * int(frame_pair_num)
        n_groups = int(frame_pair_num)
        try:
            # === 第一步：修改“分析结果.xlsx”的表头 ===
            book = load_workbook(excel_file_path)
            if 'Sheet1' not in book.sheetnames:
                raise ValueError("工作表 'Sheet1' 不存在于 '分析结果.xlsx' 中。")
            sheet = book['Sheet1']
            for col_idx, value in enumerate(full_header, 1):
                sheet.cell(row=1, column=col_idx, value=value)
            book.save(excel_file_path)
            # print(f"✅ '分析结果.xlsx' 表头已更新为 'x y u v (空)' 重复 {n_groups} 次。")
            # === 第二步：读取数据（使用 openpyxl 引擎）===
            df = pd.read_excel(excel_file_path, sheet_name='Sheet1', header=0, engine='openpyxl')
            if df.shape[0] == 0:
                raise ValueError("分析结果.xlsx 中没有数据行。")
            # === 第三步：为每一行（每个特征点）计算 u/v 的平均值 ===
            result_rows = []
            for idx, row in df.iterrows():
                u_vals = []
                v_vals = []
                x_fixed = None
                y_fixed = None
                # 遍历每一组
                for group_idx in range(n_groups):
                    start_col = group_idx * 5  # 每组5列
                    if start_col + 3 >= len(row):
                        continue
                    # 提取当前组的 x, y, u, v
                    x_val = row.iloc[start_col]
                    y_val = row.iloc[start_col + 1]
                    u_val = row.iloc[start_col + 2]
                    v_val = row.iloc[start_col + 3]
                    # 记录 u/v（跳过 NaN）
                    if pd.notna(u_val):
                        u_vals.append(u_val)
                    if pd.notna(v_val):
                        v_vals.append(v_val)
                    # 只取第一组的 x, y 作为代表（它们应该相同）
                    if group_idx == 0:
                        x_fixed = x_val
                        y_fixed = y_val
                # 计算平均值
                u_avg = pd.Series(u_vals).mean() if u_vals else None
                v_avg = pd.Series(v_vals).mean() if v_vals else None
                # 保存结果行
                result_rows.append({
                    'x': x_fixed,
                    'y': y_fixed,
                    'u': u_avg,
                    'v': v_avg
                })
            # === 第四步：保存到“平均结果.xlsx” ===
            avg_df = pd.DataFrame(result_rows)
            avg_df.to_excel(avg_excel_path, sheet_name='Sheet1', index=False)
            # print(f"✅ '平均结果.xlsx' 已创建，共 {len(avg_df)} 行数据，已保存：{avg_excel_path}")
        except Exception as e:
            print(f"❌ 操作失败：{e}")

    def _handle_apply_mask(self, dialog):  # 处理用户点击“执行掩膜”后的操作
        shape_type = dialog.shape_combo.currentText()  # 获取下拉框中当前选中的掩膜形状
        self.enable_drawing(shape_type)  # 启用绘图模式，并设置形状类型
        dialog.accept()  # 关闭对话框

    def enable_drawing(self, shape_type):  # 启用绘图模式的方法，接受形状类型（矩形或圆形）
        self.current_mask_shape = shape_type  # 保存当前掩膜形状类型
        self.right_show_panel.drawing_enabled = True  # 启用右侧图像区域的绘图模式
        self.right_show_panel.current_shape["type"] = shape_type  # 设置当前绘制形状的类型
        self.right_show_panel.setCursor(Qt.CrossCursor)  # 设置鼠标为十字准星光标（提示绘图状态）
        print(f"已启用绘图模式，当前形状：{shape_type}")  # 控制台打印调试信息

    def clear_masks(self):  # 清除所有已绘制掩膜的方法
        self.mask_shapes.clear()  # 清空掩膜图形列表
        self.right_show_panel.update()  # 刷新右侧绘图区域

    def connect_to_live_mode(self):
        """选择缓存路径并连接到IP摄像头实况模式"""
        if self.cache_path is None:
            cache_path = QFileDialog.getExistingDirectory(
                self,
                "选择缓存路径",
                QDir.homePath(),
                QFileDialog.ShowDirsOnly
            )
            if cache_path:
                self.cache_path = cache_path
            else:
                QMessageBox.warning(self, "警告", "请先选择缓存路径！")
                return
        # 如果用户选择了路径
        if self.cache_path:
            # 停止任何正在播放的视频
            self.cv_timer.stop()
            if self.cv_capture is not None:
                self.cv_capture.release()
                self.cv_capture = None
            # 显示录制视频和停止录制按钮
            self.btn_record_start.show()
            self.btn_record_stop.show()
            # 更改按钮文本为"退出实况"
            self.btn_live_mode.setText("退出实况")
            # 断开原来的连接，连接到退出实况模式的方法
            self.btn_live_mode.clicked.disconnect(self.connect_to_live_mode)
            self.btn_live_mode.clicked.connect(self.exit_the_live_mode)
            # 在单独的线程中连接摄像头
            self.connection_success = False
            self.connection_timeout = False
            # 启动连接线程
            self.connection_thread = threading.Thread(target=self.try_connect_camera)
            self.connection_thread.daemon = True
            self.connection_thread.start()
            # 启动超时检查定时器
            self.timeout_timer = QTimer()
            self.timeout_timer.setSingleShot(True)
            self.timeout_timer.timeout.connect(self.check_connection_timeout)
            self.timeout_timer.start(1000)  # 1秒超时

    def try_connect_camera(self):
        """在单独的线程中尝试连接摄像头"""
        try:
            VIDEO_URL = self.video_url
            print(f"正在尝试连接到摄像头：{VIDEO_URL}")
            cap = cv2.VideoCapture(VIDEO_URL)
            print(f"已连接到摄像头：{VIDEO_URL}")
            if self.connection_timeout:
                cap.release()
                return
            if cap.isOpened():
                self.cv_capture = cap
                self.connection_success = True
                QTimer.singleShot(0, self.on_camera_connected)
            else:
                cap.release()
                QTimer.singleShot(0, lambda: self.on_camera_connection_failed("无法打开摄像头"))
        except Exception as e:
            QTimer.singleShot(0, lambda: self.on_camera_connection_failed(str(e)))

    def check_connection_timeout(self):
        """检查连接是否超时"""
        if not self.connection_success:
            self.connection_timeout = True
            # 创建自定义对话框
            dialog = QDialog(self)
            dialog.setWindowTitle("摄像头连接")
            dialog.setModal(True)
            layout = QVBoxLayout(dialog)
            message = QLabel("开启IPcam服务器后，输入正确的摄像头IPv4地址。\n当前摄像头地址为：" + self.video_url)
            layout.addWidget(message)
            url_layout = QHBoxLayout()
            url_label = QLabel("修改地址为:")
            self.url_edit = QLineEdit(self.video_url)
            url_layout.addWidget(url_label)
            url_layout.addWidget(self.url_edit)
            layout.addLayout(url_layout)
            button_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
            button_box.accepted.connect(dialog.accept)
            button_box.rejected.connect(dialog.reject)
            layout.addWidget(button_box)
            if dialog.exec_() == QDialog.Accepted:
                # 用户点击了确定，更新URL
                new_url = self.url_edit.text().strip()
                if new_url:
                    self.video_url = new_url
                    print(f"已更新摄像头URL为: {self.video_url}")
            # 退出实况模式
            self.exit_the_live_mode()

    def on_camera_connected(self):
        """摄像头连接成功时的处理"""
        if hasattr(self, 'timeout_timer'):
            self.timeout_timer.stop()
        self.stacked_display.setCurrentIndex(1)
        self.cv_timer.start(30)  # 30毫秒一帧，大约33fps
        self.is_recording = False
        self.video_writer = None
        print("摄像头连接成功")

    def on_camera_connection_failed(self, error_msg):
        """摄像头连接失败时的处理"""
        # 停止超时定时器
        if hasattr(self, 'timeout_timer'):
            self.timeout_timer.stop()
        QMessageBox.critical(self, "连接失败", f"无法连接到摄像头：\n{error_msg}")
        self.exit_the_live_mode()

    def exit_the_live_mode(self):
        """退出实况模式"""
        # 停止定时器
        self.cv_timer.stop()
        # 释放视频捕获对象
        if self.cv_capture is not None:
            self.cv_capture.release()
            self.cv_capture = None
        if self.is_recording:
            self.stop_recording()
        self.btn_record_start.hide()
        self.btn_record_stop.hide()
        self.btn_live_mode.setText("转入实况")
        self.btn_live_mode.clicked.disconnect(self.exit_the_live_mode)
        self.btn_live_mode.clicked.connect(self.connect_to_live_mode)
        self.stacked_display.setCurrentIndex(0)

    def start_recording(self):
        """开始录制视频"""
        if self.cv_capture is not None and not self.is_recording:
            # 获取视频帧的尺寸
            self.recording_width = int(self.cv_capture.get(cv2.CAP_PROP_FRAME_WIDTH))
            self.recording_height = int(self.cv_capture.get(cv2.CAP_PROP_FRAME_HEIGHT))
            fps = self.recording_fps  # 帧率
            # 生成输出文件名（使用时间戳）
            from datetime import datetime
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_path = os.path.join(self.cache_path, f"recording_{timestamp}.mp4")
            # 使用mp4v编码器
            fourcc = cv2.VideoWriter_fourcc(*'mp4v')
            self.video_writer = cv2.VideoWriter(
                output_path, fourcc, fps,
                (self.recording_width, self.recording_height)
            )
            if self.video_writer.isOpened():
                self.is_recording = True
                self.btn_record_start.setEnabled(False)
                self.btn_record_stop.setEnabled(True)
                QMessageBox.information(self, "开始录制", f"视频将保存到: {output_path}")
            else:
                QMessageBox.critical(self, "错误", "无法创建视频文件")
                self.video_writer = None

    def stop_recording(self):
        """停止录制视频"""
        if self.is_recording and self.video_writer is not None:
            self.is_recording = False
            self.video_writer.release()
            self.video_writer = None
            self.btn_record_start.setEnabled(True)
            self.btn_record_stop.setEnabled(False)
            QMessageBox.information(self, "停止录制", "视频录制已停止")


if __name__ == '__main__':  # 判断当前模块是否为主程序入口（被直接运行而非作为模块导入）
    app = QApplication(sys.argv)  # 创建一个应用程序对象 app，传入命令行参数 sys.argv（用于支持控制台参数传递）
    ex = ImageBrowser()  # 创建 ImageBrowser 类的实例，也就是我们开发的主窗口对象
    ex.show()  # 显示主窗口，使其可见
    sys.exit(app.exec_())  # 启动 Qt 事件主循环，直到应用退出。返回退出码，传给系统




